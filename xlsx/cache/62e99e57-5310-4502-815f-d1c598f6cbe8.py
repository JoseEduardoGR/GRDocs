import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import os

# ── paleta ────────────────────────────────────────────────────────────────────
NAVY        = "1A2D5A"   # encabezado principal
NAVY_MED    = "2E4A8A"   # sub-encabezado
NAVY_LIGHT  = "D6E0F5"   # fila alterna
NAVY_ACCENT = "3A6BC4"   # bordes de totales / acento
GREEN_DARK  = "1B5E20"
GREEN_MED   = "2E7D32"
GREEN_LIGHT = "C8E6C9"
WHITE       = "FFFFFF"
YELLOW_ATT  = "FFF9C4"
BLUE_INPUT  = "0000CC"
BLACK       = "000000"
GRAY_BORDER = "BDBDBD"

def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin():
    s = Side(style="thin", color=GRAY_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium_accent():
    s = Side(style="medium", color=NAVY_ACCENT)
    return Border(left=s, right=s, top=s, bottom=s)

def font(bold=False, color=BLACK, size=11, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

MXN_FMT  = '#,##0" MXN";(#,##0" MXN");"-"'
PCT_FMT  = '0.0%'
NUM_FMT  = '#,##0'

# ── datos de ejemplo ──────────────────────────────────────────────────────────
instruments = [
    # ID, Nombre, Categoría, Marca, Modelo, Año, Estado, P.Compra, P.Venta, Stock, Ubicación
    ("INS-001","Guitarra Acústica","Cuerda","Yamaha","FG800",2022,"Nuevo",4200,7500,8,"A-01"),
    ("INS-002","Guitarra Eléctrica","Cuerda","Fender","Stratocaster",2021,"Nuevo",12500,22000,5,"A-02"),
    ("INS-003","Bajo Eléctrico","Cuerda","Gibson","SG Bass",2020,"Usado",9800,16500,3,"A-03"),
    ("INS-004","Violín 4/4","Cuerda","Stentor","Student II",2023,"Nuevo",3500,6200,10,"A-04"),
    ("INS-005","Ukulele Tenor","Cuerda","Kala","KA-T",2022,"Nuevo",1800,3200,12,"A-05"),
    ("INS-006","Guitarra Clásica","Cuerda","Cordoba","C5",2021,"Nuevo",3800,6800,7,"A-06"),
    ("INS-007","Arpa Pedal","Cuerda","Lyon & Healy","Style 23",2019,"Usado",85000,145000,1,"A-07"),
    ("INS-008","Flauta Traversa","Viento","Pearl","PF-505",2022,"Nuevo",5200,9500,6,"B-01"),
    ("INS-009","Clarinete Bb","Viento","Buffet Crampon","E13",2021,"Nuevo",8900,15800,4,"B-02"),
    ("INS-010","Saxofón Alto","Viento","Yamaha","YAS-280",2022,"Nuevo",14500,26000,3,"B-03"),
    ("INS-011","Trompeta Bb","Viento","Bach","TR300H2",2020,"Nuevo",9200,16500,5,"B-04"),
    ("INS-012","Trombón Tenor","Viento","Conn-Selmer","88H",2019,"Usado",18000,32000,2,"B-05"),
    ("INS-013","Oboe","Viento","Fox","Renard 330",2021,"Nuevo",28000,48000,2,"B-06"),
    ("INS-014","Corno Francés","Viento","Holton","H179",2020,"Reparación",22000,40000,1,"B-07"),
    ("INS-015","Batería Completa","Percusión","Pearl","Export",2022,"Nuevo",18500,32000,4,"C-01"),
    ("INS-016","Cajón Flamenco","Percusión","LP","LPA1331",2023,"Nuevo",2200,4200,9,"C-02"),
    ("INS-017","Marimba 4.3 Oct","Percusión","Adams","Soloist",2021,"Nuevo",45000,78000,1,"C-03"),
    ("INS-018","Xilófono","Percusión","Sonor","GS Play",2022,"Nuevo",8500,15000,3,"C-04"),
    ("INS-019","Bongos","Percusión","Meinl","Headliner",2023,"Nuevo",1400,2800,15,"C-05"),
    ("INS-020","Timbales Orq.","Percusión","Ludwig","LKS625FG",2020,"Usado",35000,60000,2,"C-06"),
]

months = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
ventas_data = [
    # mes, unidades_vendidas, ingresos
    ("Ene 2024", 18, 145000),
    ("Feb 2024", 22, 178000),
    ("Mar 2024", 31, 265000),
    ("Abr 2024", 27, 218000),
    ("May 2024", 35, 312000),
    ("Jun 2024", 42, 380000),
    ("Jul 2024", 38, 345000),
    ("Ago 2024", 29, 258000),
    ("Sep 2024", 33, 295000),
    ("Oct 2024", 45, 415000),
    ("Nov 2024", 52, 480000),
    ("Dic 2024", 61, 570000),
]

# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── HOJA 1: INVENTARIO ────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Inventario"

# Título principal
ws1.merge_cells("A1:K1")
c = ws1["A1"]
c.value = "🎵  CONTROL DE INVENTARIO DE INSTRUMENTOS MUSICALES"
c.fill   = solid(NAVY)
c.font   = font(bold=True, color=WHITE, size=16)
c.alignment = align("center")
ws1.row_dimensions[1].height = 40

# Sub-título
ws1.merge_cells("A2:K2")
c = ws1["A2"]
c.value = "Gestión y seguimiento de activos musicales"
c.fill  = solid(NAVY_MED)
c.font  = font(color=WHITE, size=11, italic=True)
c.alignment = align("center")
ws1.row_dimensions[2].height = 22

# Fila vacía de separación
ws1.row_dimensions[3].height = 8

# Encabezados de columna
headers = [
    "ID","Nombre del Instrumento","Categoría","Marca","Modelo",
    "Año Fab.","Estado","Precio Compra","Precio Venta","Stock Disp.","Ubicación"
]
for col_idx, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col_idx, value=h)
    c.fill      = solid(NAVY)
    c.font      = font(bold=True, color=WHITE, size=10)
    c.alignment = align("center")
    c.border    = border_thin()
ws1.row_dimensions[4].height = 28

# Datos
for row_idx, inst in enumerate(instruments, 5):
    is_alt = (row_idx % 2 == 0)
    bg = NAVY_LIGHT if is_alt else WHITE
    needs_attention = inst[9] < 5  # stock < 5

    for col_idx, val in enumerate(inst, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        c.border    = border_thin()
        c.alignment = align("center" if col_idx not in (2,4,5) else "left")
        row_bg = YELLOW_ATT if needs_attention else bg

        if col_idx in (8, 9):           # precios
            c.fill         = solid(row_bg)
            c.font         = font(color=BLUE_INPUT, size=10)
            c.number_format = MXN_FMT
        elif col_idx == 10:             # stock
            c.fill = solid(row_bg)
            c.font = font(color=BLUE_INPUT, size=10, bold=needs_attention)
            c.number_format = NUM_FMT
        else:
            c.fill = solid(row_bg)
            c.font = font(color=BLACK, size=10)
    ws1.row_dimensions[row_idx].height = 20

# Fila totales
total_row = len(instruments) + 5
ws1.merge_cells(f"A{total_row}:G{total_row}")
c = ws1[f"A{total_row}"]
c.value     = "TOTALES"
c.fill      = solid(NAVY_MED)
c.font      = font(bold=True, color=WHITE)
c.alignment = align("right")
c.border    = border_medium_accent()
ws1.row_dimensions[total_row].height = 24

# Suma precios compra
data_start = 5
data_end   = len(instruments) + 4

for col_letter, col_idx in [("H", 8), ("I", 9), ("J", 10)]:
    c = ws1.cell(row=total_row, column=col_idx)
    c.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
    c.fill  = solid(NAVY_MED)
    c.font  = font(bold=True, color=WHITE)
    c.alignment = align("center")
    c.border    = border_medium_accent()
    if col_idx in (8, 9):
        c.number_format = MXN_FMT
    else:
        c.number_format = NUM_FMT

ws1.cell(row=total_row, column=11).fill   = solid(NAVY_MED)
ws1.cell(row=total_row, column=11).border = border_medium_accent()

# Anchos de columna
col_widths = [10,28,14,14,16,10,12,16,16,12,12]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Congelar paneles
ws1.freeze_panes = "A5"

# ── HOJA 2: RESUMEN ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Resumen")

ws2.merge_cells("A1:H1")
c = ws2["A1"]
c.value = "📊  RESUMEN EJECUTIVO — INVENTARIO MUSICAL"
c.fill  = solid(NAVY)
c.font  = font(bold=True, color=WHITE, size=15)
c.alignment = align("center")
ws2.row_dimensions[1].height = 38

ws2.merge_cells("A2:H2")
c = ws2["A2"]
c.value = "Indicadores clave de gestión de inventario"
c.fill  = solid(NAVY_MED)
c.font  = font(color=WHITE, italic=True)
c.alignment = align("center")
ws2.row_dimensions[2].height = 20

# ── Sección A: Totales por Categoría ──────────────────────────────────────────
ws2.merge_cells("A4:D4")
c = ws2["A4"]
c.value = "INVENTARIO POR CATEGORÍA"
c.fill  = solid(NAVY)
c.font  = font(bold=True, color=WHITE)
c.alignment = align("center")
ws2.row_dimensions[4].height = 24

cat_headers = ["Categoría","Cantidad (Items)","Stock Total","Valor Inventario"]
for ci, h in enumerate(cat_headers, 1):
    c = ws2.cell(row=5, column=ci, value=h)
    c.fill  = solid(NAVY_MED)
    c.font  = font(bold=True, color=WHITE, size=10)
    c.alignment = align("center")
    c.border = border_thin()
ws2.row_dimensions[5].height = 22

categories = ["Cuerda","Viento","Percusión"]
for ri, cat in enumerate(categories, 6):
    bg = NAVY_LIGHT if ri % 2 == 0 else WHITE
    # Categoría
    c = ws2.cell(row=ri, column=1, value=cat)
    c.fill = solid(bg); c.font = font(bold=True); c.alignment = align("center"); c.border = border_thin()

    # Cantidad (COUNTIF en hoja Inventario)
    c = ws2.cell(row=ri, column=2)
    c.value = f'=COUNTIF(Inventario!C{data_start}:C{data_end},A{ri})'
    c.fill = solid(bg); c.font = font(color=BLACK); c.alignment = align("center"); c.border = border_thin()
    c.number_format = NUM_FMT

    # Stock total (SUMIF)
    c = ws2.cell(row=ri, column=3)
    c.value = f'=SUMIF(Inventario!C{data_start}:C{data_end},A{ri},Inventario!J{data_start}:J{data_end})'
    c.fill = solid(bg); c.font = font(color=BLACK); c.alignment = align("center"); c.border = border_thin()
    c.number_format = NUM_FMT

    # Valor inventario (SUMIF precio venta * stock)
    c = ws2.cell(row=ri, column=4)
    c.value = f'=SUMPRODUCT((Inventario!C{data_start}:C{data_end}=A{ri})*(Inventario!I{data_start}:I{data_end})*(Inventario!J{data_start}:J{data_end}))'
    c.fill = solid(bg); c.font = font(color=GREEN_DARK, bold=True); c.alignment = align("center"); c.border = border_thin()
    c.number_format = MXN_FMT
    ws2.row_dimensions[ri].height = 20

# Fila total categorías
tr = 9
ws2.cell(row=tr, column=1, value="TOTAL GENERAL").font = font(bold=True, color=WHITE)
ws2.cell(row=tr, column=1).fill = solid(NAVY_MED); ws2.cell(row=tr, column=1).alignment = align("center"); ws2.cell(row=tr, column=1).border = border_medium_accent()

for ci2 in range(2, 5):
    c = ws2.cell(row=tr, column=ci2)
    col_l = get_column_letter(ci2)
    c.value = f"=SUM({col_l}6:{col_l}8)"
    c.fill  = solid(NAVY_MED)
    c.font  = font(bold=True, color=WHITE)
    c.alignment = align("center")
    c.border = border_medium_accent()
    c.number_format = MXN_FMT if ci2 == 4 else NUM_FMT
ws2.row_dimensions[tr].height = 24

# ── Gráfico de barras por categoría ───────────────────────────────────────────
bar_chart = BarChart()
bar_chart.type  = "col"
bar_chart.style = 10
bar_chart.title = "Inventario por Categoría"
bar_chart.y_axis.title = "Cantidad"
bar_chart.x_axis.title = "Categoría"
bar_chart.shape = 4
bar_chart.width  = 14
bar_chart.height = 10

data_ref = Reference(ws2, min_col=2, max_col=3, min_row=5, max_row=8)
cats_ref = Reference(ws2, min_col=1, min_row=6, max_row=8)
bar_chart.add_data(data_ref, titles_from_data=True)
bar_chart.set_categories(cats_ref)
ws2.add_chart(bar_chart, "F4")

# ── Sección B: KPIs ───────────────────────────────────────────────────────────
ws2.merge_cells("A11:D11")
c = ws2["A11"]
c.value = "INDICADORES CLAVE DE INVENTARIO"
c.fill  = solid(NAVY); c.font = font(bold=True, color=WHITE); c.alignment = align("center")
ws2.row_dimensions[11].height = 24

kpi_labels = [
    ("Valor Total Inventario (Costo)",   f'=SUMPRODUCT(Inventario!H{data_start}:H{data_end},Inventario!J{data_start}:J{data_end})', MXN_FMT, False),
    ("Valor Total Inventario (Venta)",   f'=SUMPRODUCT(Inventario!I{data_start}:I{data_end},Inventario!J{data_start}:J{data_end})', MXN_FMT, False),
    ("Margen Promedio",                  f'=IFERROR((SUMPRODUCT(Inventario!I{data_start}:I{data_end},Inventario!J{data_start}:J{data_end})-SUMPRODUCT(Inventario!H{data_start}:H{data_end},Inventario!J{data_start}:J{data_end}))/SUMPRODUCT(Inventario!H{data_start}:H{data_end},Inventario!J{data_start}:J{data_end}),0)', PCT_FMT, False),
    ("Total Instrumentos en Catálogo",   f'=COUNTA(Inventario!A{data_start}:A{data_end})', NUM_FMT, False),
    ("Instrumentos con Stock < 5",       f'=COUNTIF(Inventario!J{data_start}:J{data_end},"<5")', NUM_FMT, True),
    ("Instrumentos en Reparación",       f'=COUNTIF(Inventario!G{data_start}:G{data_end},"Reparación")', NUM_FMT, True),
]

for ri, (lbl, formula, fmt, attention) in enumerate(kpi_labels, 12):
    bg = YELLOW_ATT if attention else (NAVY_LIGHT if ri % 2 == 0 else WHITE)
    ws2.merge_cells(f"A{ri}:C{ri}")
    c = ws2[f"A{ri}"]
    c.value = lbl; c.fill = solid(bg); c.font = font(bold=True); c.alignment = align("left"); c.border = border_thin()

    c2 = ws2.cell(row=ri, column=4, value=formula)
    c2.fill = solid(bg); c2.font = font(color=GREEN_DARK, bold=True); c2.alignment = align("center"); c2.border = border_thin()
    c2.number_format = fmt
    ws2.row_dimensions[ri].height = 22

# ── Sección C: Top 5 más caros ────────────────────────────────────────────────
ws2.merge_cells("A19:D19")
c = ws2["A19"]
c.value = "TOP 5 — INSTRUMENTOS MÁS CAROS (PRECIO VENTA)"
c.fill  = solid(NAVY); c.font = font(bold=True, color=WHITE); c.alignment = align("center")
ws2.row_dimensions[19].height = 24

top5_h = ["#","Instrumento","Categoría","Precio Venta"]
for ci, h in enumerate(top5_h, 1):
    c = ws2.cell(row=20, column=ci, value=h)
    c.fill = solid(NAVY_MED); c.font = font(bold=True, color=WHITE, size=10)
    c.alignment = align("center"); c.border = border_thin()
ws2.row_dimensions[20].height = 22

for rank in range(1, 6):
    ri = 20 + rank
    bg = NAVY_LIGHT if rank % 2 == 0 else WHITE
    ws2.cell(row=ri, column=1, value=rank).fill = solid(bg)
    ws2.cell(row=ri, column=1).font = font(bold=True); ws2.cell(row=ri, column=1).alignment = align("center"); ws2.cell(row=ri, column=1).border = border_thin()

    # LARGE + INDEX/MATCH para nombre
    c2 = ws2.cell(row=ri, column=2)
    c2.value = f'=IFERROR(INDEX(Inventario!B{data_start}:B{data_end},MATCH(LARGE(Inventario!I{data_start}:I{data_end},{rank}),Inventario!I{data_start}:I{data_end},0)),"N/A")'
    c2.fill = solid(bg); c2.font = font(color=BLACK); c2.alignment = align("left"); c2.border = border_thin()

    c3 = ws2.cell(row=ri, column=3)
    c3.value = f'=IFERROR(INDEX(Inventario!C{data_start}:C{data_end},MATCH(LARGE(Inventario!I{data_start}:I{data_end},{rank}),Inventario!I{data_start}:I{data_end},0)),"N/A")'
    c3.fill = solid(bg); c3.font = font(color=BLACK); c3.alignment = align("center"); c3.border = border_thin()

    c4 = ws2.cell(row=ri, column=4)
    c4.value = f'=IFERROR(LARGE(Inventario!I{data_start}:I{data_end},{rank}),0)'
    c4.fill = solid(bg); c4.font = font(color=GREEN_DARK, bold=True); c4.alignment = align("center"); c4.border = border_thin()
    c4.number_format = MXN_FMT
    ws2.row_dimensions[ri].height = 20

# Anchos hoja 2
for i, w in enumerate([18,28,18,20,2,2], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.column_dimensions["F"].width = 2
ws2.column_dimensions["G"].width = 14

ws2.freeze_panes = "A3"

# ── HOJA 3: VENTAS MENSUALES ──────────────────────────────────────────────────
ws3 = wb.create_sheet("Ventas Mensuales")

ws3.merge_cells("A1:F1")
c = ws3["A1"]
c.value = "📈  VENTAS MENSUALES — EJERCICIO 2024"
c.fill  = solid(NAVY)
c.font  = font(bold=True, color=WHITE, size=15)
c.alignment = align("center")
ws3.row_dimensions[1].height = 38

ws3.merge_cells("A2:F2")
c = ws3["A2"]
c.value = "Análisis de tendencia de ingresos por mes"
c.fill  = solid(NAVY_MED); c.font = font(color=WHITE, italic=True); c.alignment = align("center")
ws3.row_dimensions[2].height = 20

# Encabezados tabla ventas
v_headers = ["Mes","Unidades Vendidas","Ingresos Brutos","Promedio x Unidad","Var. vs Mes Ant.","% Variación"]
for ci, h in enumerate(v_headers, 1):
    c = ws3.cell(row=4, column=ci, value=h)
    c.fill  = solid(NAVY)
    c.font  = font(bold=True, color=WHITE, size=10)
    c.alignment = align("center", wrap=True)
    c.border = border_thin()
ws3.row_dimensions[4].height = 32

for ri, (mes, unidades, ingresos) in enumerate(ventas_data, 5):
    bg = NAVY_LIGHT if ri % 2 == 0 else WHITE

    c = ws3.cell(row=ri, column=1, value=mes)
    c.fill = solid(bg); c.font = font(bold=True); c.alignment = align("center"); c.border = border_thin()

    c = ws3.cell(row=ri, column=2, value=unidades)
    c.fill = solid(bg); c.font = font(color=BLUE_INPUT); c.alignment = align("center"); c.border = border_thin()
    c.number_format = NUM_FMT

    c = ws3.cell(row=ri, column=3, value=ingresos)
    c.fill = solid(bg); c.font = font(color=BLUE_INPUT); c.alignment = align("center"); c.border = border_thin()
    c.number_format = MXN_FMT

    # Promedio por unidad
    c = ws3.cell(row=ri, column=4)
    c.value = f"=IFERROR(C{ri}/B{ri},0)"
    c.fill = solid(bg); c.font = font(color=BLACK); c.alignment = align("center"); c.border = border_thin()
    c.number_format = MXN_FMT

    # Variación vs mes anterior
    if ri == 5:
        ws3.cell(row=ri, column=5, value="-").fill = solid(bg)
        ws3.cell(row=ri, column=5).font = font(color=BLACK); ws3.cell(row=ri, column=5).alignment = align("center"); ws3.cell(row=ri, column=5).border = border_thin()
        ws3.cell(row=ri, column=6, value="-").fill = solid(bg)
        ws3.cell(row=ri, column=6).font = font(color=BLACK); ws3.cell(row=ri, column=6).alignment = align("center"); ws3.cell(row=ri, column=6).border = border_thin()
    else:
        c5 = ws3.cell(row=ri, column=5)
        c5.value = f"=C{ri}-C{ri-1}"
        c5.fill = solid(bg); c5.font = font(color=BLACK); c5.alignment = align("center"); c5.border = border_thin()
        c5.number_format = MXN_FMT

        c6 = ws3.cell(row=ri, column=6)
        c6.value = f"=IFERROR((C{ri}-C{ri-1})/C{ri-1},0)"
        c6.fill = solid(bg); c6.font = font(color=BLACK); c6.alignment = align("center"); c6.border = border_thin()
        c6.number_format = PCT_FMT

    ws3.row_dimensions[ri].height = 22

# Fila de totales y promedios
total_r = len(ventas_data) + 5
ws3.cell(row=total_r, column=1, value="TOTAL / PROMEDIO").fill = solid(NAVY_MED)
ws3.cell(row=total_r, column=1).font = font(bold=True, color=WHITE); ws3.cell(row=total_r, column=1).alignment = align("center"); ws3.cell(row=total_r, column=1).border = border_medium_accent()

for ci, formula, fmt in [
    (2, f"=SUM(B5:B{total_r-1})", NUM_FMT),
    (3, f"=SUM(C5:C{total_r-1})", MXN_FMT),
    (4, f"=AVERAGE(D5:D{total_r-1})", MXN_FMT),
]:
    c = ws3.cell(row=total_r, column=ci, value=formula)
    c.fill = solid(NAVY_MED); c.font = font(bold=True, color=WHITE); c.alignment = align("center")
    c.border = border_medium_accent(); c.number_format = fmt

for ci in (5, 6):
    ws3.cell(row=total_r, column=ci).fill = solid(NAVY_MED)
    ws3.cell(row=total_r, column=ci).border = border_medium_accent()

ws3.row_dimensions[total_r].height = 26

# ── Bloque promedio mensual destacado ─────────────────────────────────────────
avg_r = total_r + 2
ws3.merge_cells(f"A{avg_r}:C{avg_r}")
c = ws3[f"A{avg_r}"]
c.value = "Promedio Mensual de Ingresos"
c.fill = solid(GREEN_LIGHT); c.font = font(bold=True, color=GREEN_DARK, size=12); c.alignment = align("center"); c.border = border_thin()

c4 = ws3.cell(row=avg_r, column=4)
c4.value = f"=AVERAGE(C5:C{total_r-1})"
c4.fill = solid(GREEN_LIGHT); c4.font = font(bold=True, color=GREEN_DARK, size=12); c4.alignment = align("center"); c4.border = border_thin()
c4.number_format = MXN_FMT
ws3.row_dimensions[avg_r].height = 28

# ── Gráfico de línea ──────────────────────────────────────────────────────────
line_chart = LineChart()
line_chart.title  = "Tendencia de Ingresos Mensuales 2024"
line_chart.y_axis.title = "Ingresos (MXN)"
line_chart.x_axis.title = "Mes"
line_chart.style  = 10
line_chart.width  = 20
line_chart.height = 12

data_ref2 = Reference(ws3, min_col=3, max_col=3, min_row=4, max_row=total_r-1)
cats_ref2 = Reference(ws3, min_col=1, min_row=5, max_row=total_r-1)
line_chart.add_data(data_ref2, titles_from_data=True)
line_chart.set_categories(cats_ref2)

series = line_chart.series[0]
series.graphicalProperties.line.solidFill   = NAVY_ACCENT
series.graphicalProperties.line.width       = 28000
series.marker.symbol                        = "circle"
series.marker.size                          = 7
series.marker.graphicalProperties.fgColor   = GREEN_MED
series.marker.graphicalProperties.solidFill = GREEN_MED

ws3.add_chart(line_chart, "A20")

# Anchos hoja 3
col3_widths = [14, 18, 20, 20, 18, 14]
for i, w in enumerate(col3_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = "A5"

# ── Guardar ───────────────────────────────────────────────────────────────────
output_path = "xlsx/cache/output/c59904ce-d2c1-450a-8fa8-98d3e3d0bee9.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)

size_kb = os.path.getsize(output_path) / 1024
print(f"✅ Archivo generado exitosamente: {output_path}")
print(f"   Tamaño: {size_kb:.1f} KB")
print(f"   Hojas: Inventario ({len(instruments)} instrumentos) | Resumen | Ventas Mensuales ({len(ventas_data)} meses)")